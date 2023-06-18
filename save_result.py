from openpyxl import Workbook
from openpyxl.styles import Font


class SaveResult:
    def __init__(self, good_dict):
        self.good_dict = good_dict

        self.colums = ['name_post', 'link', 'views', 'date', 'name_them', 'name_author', 'post_text', 'likes_post',
                       'author_comment', 'text_comment', 'like_comment', 'time_comment']

        self.comment_colums = 9

    def create_title(self, ws):
        for count, col in enumerate(self.colums):
            ws.cell(row=1, column=count + 1).value = col
            ws.cell(row=1, column=count + 1).font = Font(bold=True)

    def merge_row(self, ws, count_comments, count_def):

        for count in range(self.comment_colums - 1):
            count = count + 1
            ws.merge_cells(start_column=count, start_row=count_def,
                           end_column=count, end_row=count_def + count_comments - 1)

        return True

    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = post['name_post']
        ws.cell(row=count_def, column=2).value = post['link']
        ws.cell(row=count_def, column=3).value = post['views_post']
        ws.cell(row=count_def, column=4).value = post['date_post']
        ws.cell(row=count_def, column=5).value = post['name_them']
        ws.cell(row=count_def, column=6).value = post['name_author']
        ws.cell(row=count_def, column=7).value = post['text_post']
        ws.cell(row=count_def, column=8).value = int(post['like'])


        if len(post['comments']) == 0:
            return True

        for count_com, comment in enumerate(post['comments']):
            ws.cell(row=count_def + count_com, column=9).value = comment['name_comment']
            ws.cell(row=count_def + count_com, column=10).value = comment['text_comment']

            if comment['like'] == '':
                comment['like'] = 0

            ws.cell(row=count_def + count_com, column=11).value = int(comment['like'])
            ws.cell(row=count_def + count_com, column=12).value = comment['time_comment']

        return True

    def itter_rows(self, ws):
        count_def = 2  # с двойки т.к. в 1 строке уже загол
        for count_post, post in enumerate(self.good_dict):
            count_comments = len(post['comments'])

            if count_comments > 1:
                response = self.merge_row(ws, count_comments, count_def)

            write_data = self.write_data(ws, count_def, post)

            if count_comments > 1:
                count_def = count_def + count_comments
            else:
                count_def += 1


    def one_sheet(self, ws):

        response = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        filename = f'{filename}.xlsx'

        wb.save(filename)

        return filename
